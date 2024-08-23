import openpyxl
from datetime import datetime, timedelta

def setAvd(sheet_names):
    for i in range(0,len(sheet_names)):
        print(i,' ',sheet_names[i])
    return sheet_names[int(input('Vilken avdlening: '))]

def setKollega(schemaTab):
    kollegor = []
    for i in range(1, 200):
        cell = 'B'+str(i)
        if schemaTab[cell].value !='' and schemaTab[cell].value != None and schemaTab[cell].value != 'Namn':
            kollegor.append(schemaTab[cell].value)
    for i in range(0, len(kollegor)):
        print(str(i)+' '+kollegor[i])
    return kollegor[int(input('Vilken kollega: '))]

def setCellValue(schemaTab,letter, num):
    value = str(schemaTab[letter+str(num)].value.hour)+','+str(schemaTab[letter+str(num)].value.minute)
    return value

def setRowValues(schemaTab, rowSchema, mallTab, rowMall, rast, letters):
    startLetter = letters[0]
    endLetter = letters[1]
    egenPlanStart = letters[2]
    egenPlanSlut = letters[3]
    rasten = rast
    
    if schemaTab[egenPlanStart+str(rowSchema)].value  is not None and schemaTab[egenPlanSlut+str(rowSchema)].value is not None:
        if schemaTab[endLetter+str(rowSchema)].value != schemaTab[egenPlanStart+str(rowSchema)].value:
            time1 = datetime.combine(datetime.today(), schemaTab[endLetter+str(rowSchema)].value)
            time2 = datetime.combine(datetime.today(), schemaTab[egenPlanStart+str(rowSchema)].value)
            rastDiff = time2-time1
            rastDiff_in_minutes = int(rastDiff.total_seconds() / 60)
            print('EP:',time1,'-',time2,'=',rastDiff_in_minutes)
            rasten = str(int(rasten) + rastDiff_in_minutes)
        endLetter = egenPlanSlut
    
    if len(letters) == 6:
        moteStart = letters[4]
        moteSlut = letters[5]
        if schemaTab[moteStart+str(rowSchema)].value  is not None and schemaTab[moteSlut+str(rowSchema)].value is not None:
            if schemaTab[moteStart+str(rowSchema)].value  is not None and schemaTab[moteSlut+str(rowSchema)].value is not None:
                time1 = datetime.combine(datetime.today(), schemaTab[endLetter+str(rowSchema)].value)
                time2 = datetime.combine(datetime.today(), schemaTab[moteStart+str(rowSchema)].value)
                rastDiff = time2-time1
                rastDiff_in_minutes = int(rastDiff.total_seconds() / 60)
                print('Möte::',time1,'-',time2,'=',rastDiff_in_minutes)
                rasten = str(int(rasten) + rastDiff_in_minutes)
            endLetter = moteSlut

    mallTab['C'+str(rowMall)] = setCellValue(schemaTab, startLetter, rowSchema)
    mallTab['D'+str(rowMall)] = setCellValue(schemaTab, endLetter, rowSchema)
    mallTab['E'+str(rowMall)] = rasten

def main():
    schema = openpyxl.load_workbook('Schema.xlsx')
    mall = openpyxl.load_workbook('mall.xlsx')

    sheet_names = [sheet for sheet in schema.sheetnames if schema[sheet].sheet_state == 'visible']

    tabName = setAvd(sheet_names)
    schemaTab = schema[tabName]
    namn = setKollega(schemaTab)
    startdatum = input('Periodens start datum (xxxx-xx-xx): ')
    numVeckor = input('Hur många veckor i perioden: ')
    rast = input('Hur lång rast (Minuter): ')
    numDagar = int(numVeckor)*7
    mallTab = mall['Schema mall']

    startPosSchema = ''

    for x in range(1,200):
        cell = 'B'+str(x)
        val = schemaTab[cell].value
        if val == namn:
            startPosSchema = str(x)
            break

    mallTab['C8'] = int(numVeckor)

    i = 0
    rowSchema = int(startPosSchema)
    rowMall = 14

    while i < numDagar:
        dagar = [['J','K','O','P'],['R', 'S','W','X','Z','AA'],['AC', 'AD','AH','AI','AK','AL'],['AN', 'AO','AS','AT','AV','AW'],['AY', 'AZ','BD','BE']]
        for x in dagar:
            setRowValues(schemaTab, rowSchema, mallTab, rowMall, rast, x)
            rowMall +=1
        #Lördag och Söndag
        rowMall +=2

        rowSchema +=1
        i+=7

    mallTab['C8'] = int(numVeckor)
    mallTab['B12'] = startdatum

    print('Tider inlagd!')

    fulltNamn = input('Ange kollegans förnamn och efternamn: ')
    personnummer = input('Ange kollegans personnummer: ')

    mallTab['C4'] = fulltNamn
    mallTab['C5'] = personnummer

    newFile = fulltNamn+'.xlsx'

    mall.save(newFile)

    print('Klar!!!')


main()
